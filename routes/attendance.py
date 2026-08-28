
from flask import (
    Blueprint,
    render_template,
    session,
    redirect,
    url_for,
    request,
    flash,
    send_file
)

from extensions import mysql

from openpyxl import Workbook

from io import BytesIO

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle
)

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch


attendance = Blueprint(
    "attendance",
    __name__,
    url_prefix="/attendance"
)


# ==========================================================
# HELPER
# Generate next attendance_sessions.id
# ==========================================================

def get_next_session_id(cursor):
    """
    attendance_sessions.id is currently NOT AUTO_INCREMENT.

    Therefore we manually generate the next ID.
    """

    cursor.execute("""
        SELECT COALESCE(MAX(id), 0) + 1
        FROM attendance_sessions
    """)

    result = cursor.fetchone()

    if not result or result[0] is None:
        return 1

    return int(result[0])


# ==========================================================
# Attendance Session List
# ==========================================================

@attendance.route("/")
def index():

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    cursor = mysql.connection.cursor()

    try:

        cursor.execute("""
            SELECT
                attendance_sessions.id,
                subjects.subject_name,
                teachers.full_name,
                attendance_sessions.session_date,
                attendance_sessions.start_time,
                attendance_sessions.end_time,
                attendance_sessions.session_status

            FROM attendance_sessions

            INNER JOIN subjects
                ON attendance_sessions.subject_id = subjects.id

            INNER JOIN teachers
                ON attendance_sessions.teacher_id = teachers.id

            ORDER BY attendance_sessions.id DESC
        """)

        sessions = cursor.fetchall()

    except Exception as e:

        print(
            "ATTENDANCE SESSION LIST ERROR:",
            str(e)
        )

        flash(
            f"Unable to load attendance sessions: {e}",
            "danger"
        )

        sessions = []

    finally:

        cursor.close()

    return render_template(
        "admin/attendance.html",
        sessions=sessions
    )


# ==========================================================
# Add Attendance Session
# ==========================================================

@attendance.route(
    "/add",
    methods=["GET", "POST"]
)
def add_session():

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    cursor = mysql.connection.cursor()

    try:

        # ==================================================
        # POST
        # ==================================================

        if request.method == "POST":

            subject_id = request.form.get(
                "subject_id",
                ""
            ).strip()

            teacher_id = request.form.get(
                "teacher_id",
                ""
            ).strip()

            session_date = request.form.get(
                "session_date",
                ""
            ).strip()

            start_time = request.form.get(
                "start_time",
                ""
            ).strip()

            end_time = request.form.get(
                "end_time",
                ""
            ).strip()

            session_status = request.form.get(
                "session_status",
                "OPEN"
            ).strip().upper()

            # ----------------------------------------------
            # Validation
            # ----------------------------------------------

            if not subject_id:
                flash(
                    "Please select a subject.",
                    "warning"
                )

                return redirect(
                    url_for("attendance.add_session")
                )

            if not teacher_id:
                flash(
                    "Please select a teacher.",
                    "warning"
                )

                return redirect(
                    url_for("attendance.add_session")
                )

            if not session_date:
                flash(
                    "Please select session date.",
                    "warning"
                )

                return redirect(
                    url_for("attendance.add_session")
                )

            if not start_time:
                flash(
                    "Please select start time.",
                    "warning"
                )

                return redirect(
                    url_for("attendance.add_session")
                )

            if session_status not in (
                "OPEN",
                "CLOSED"
            ):
                session_status = "OPEN"

            # ----------------------------------------------
            # Verify Subject
            # ----------------------------------------------

            cursor.execute("""
                SELECT id
                FROM subjects
                WHERE id=%s
                LIMIT 1
            """, (
                subject_id,
            ))

            subject_exists = cursor.fetchone()

            if not subject_exists:

                flash(
                    "Selected subject does not exist.",
                    "danger"
                )

                return redirect(
                    url_for("attendance.add_session")
                )

            # ----------------------------------------------
            # Verify Teacher
            # ----------------------------------------------

            cursor.execute("""
                SELECT id
                FROM teachers
                WHERE id=%s
                LIMIT 1
            """, (
                teacher_id,
            ))

            teacher_exists = cursor.fetchone()

            if not teacher_exists:

                flash(
                    "Selected teacher does not exist.",
                    "danger"
                )

                return redirect(
                    url_for("attendance.add_session")
                )

            # ----------------------------------------------
            # Generate ID
            #
            # IMPORTANT:
            # id is NOT AUTO_INCREMENT in TiDB.
            # ----------------------------------------------

            new_session_id = get_next_session_id(
                cursor
            )

            # ----------------------------------------------
            # Insert Attendance Session
            # ----------------------------------------------

            cursor.execute("""
                INSERT INTO attendance_sessions
                (
                    id,
                    subject_id,
                    teacher_id,
                    session_date,
                    start_time,
                    end_time,
                    session_status
                )
                VALUES
                (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s
                )
            """, (
                new_session_id,
                subject_id,
                teacher_id,
                session_date,
                start_time,
                end_time if end_time else None,
                session_status
            ))

            mysql.connection.commit()

            flash(
                "Attendance Session created successfully!",
                "success"
            )

            return redirect(
                url_for("attendance.index")
            )

        # ==================================================
        # GET
        # ==================================================

        cursor.execute("""
            SELECT
                id,
                subject_name
            FROM subjects
            ORDER BY subject_name
        """)

        subjects = cursor.fetchall()

        cursor.execute("""
            SELECT
                id,
                full_name
            FROM teachers
            ORDER BY full_name
        """)

        teachers = cursor.fetchall()

        return render_template(
            "admin/add_attendance.html",
            subjects=subjects,
            teachers=teachers
        )

    except Exception as e:

        mysql.connection.rollback()

        print(
            "ADMIN ATTENDANCE SESSION CREATE ERROR:",
            str(e)
        )

        flash(
            f"Error creating attendance session: {e}",
            "danger"
        )

        return redirect(
            url_for("attendance.add_session")
        )

    finally:

        cursor.close()


# ==========================================================
# Edit Attendance Session
# ==========================================================

@attendance.route(
    "/edit/<int:id>",
    methods=["GET", "POST"]
)
def edit_session(id):

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    cursor = mysql.connection.cursor()

    try:

        # ==================================================
        # POST
        # ==================================================

        if request.method == "POST":

            subject_id = request.form.get(
                "subject_id",
                ""
            ).strip()

            teacher_id = request.form.get(
                "teacher_id",
                ""
            ).strip()

            session_date = request.form.get(
                "session_date",
                ""
            ).strip()

            start_time = request.form.get(
                "start_time",
                ""
            ).strip()

            end_time = request.form.get(
                "end_time",
                ""
            ).strip()

            session_status = request.form.get(
                "session_status",
                "OPEN"
            ).strip().upper()

            # ----------------------------------------------
            # Validation
            # ----------------------------------------------

            if not subject_id or not teacher_id:
                flash(
                    "Subject and teacher are required.",
                    "warning"
                )

                return redirect(
                    url_for(
                        "attendance.edit_session",
                        id=id
                    )
                )

            if not session_date:
                flash(
                    "Session date is required.",
                    "warning"
                )

                return redirect(
                    url_for(
                        "attendance.edit_session",
                        id=id
                    )
                )

            if not start_time:
                flash(
                    "Start time is required.",
                    "warning"
                )

                return redirect(
                    url_for(
                        "attendance.edit_session",
                        id=id
                    )
                )

            if session_status not in (
                "OPEN",
                "CLOSED"
            ):
                session_status = "OPEN"

            # ----------------------------------------------
            # Check Session
            # ----------------------------------------------

            cursor.execute("""
                SELECT id
                FROM attendance_sessions
                WHERE id=%s
                LIMIT 1
            """, (
                id,
            ))

            existing = cursor.fetchone()

            if not existing:

                flash(
                    "Attendance session not found.",
                    "danger"
                )

                return redirect(
                    url_for("attendance.index")
                )

            # ----------------------------------------------
            # Verify Subject
            # ----------------------------------------------

            cursor.execute("""
                SELECT id
                FROM subjects
                WHERE id=%s
                LIMIT 1
            """, (
                subject_id,
            ))

            if not cursor.fetchone():

                flash(
                    "Selected subject does not exist.",
                    "danger"
                )

                return redirect(
                    url_for(
                        "attendance.edit_session",
                        id=id
                    )
                )

            # ----------------------------------------------
            # Verify Teacher
            # ----------------------------------------------

            cursor.execute("""
                SELECT id
                FROM teachers
                WHERE id=%s
                LIMIT 1
            """, (
                teacher_id,
            ))

            if not cursor.fetchone():

                flash(
                    "Selected teacher does not exist.",
                    "danger"
                )

                return redirect(
                    url_for(
                        "attendance.edit_session",
                        id=id
                    )
                )

            # ----------------------------------------------
            # Update
            # ----------------------------------------------

            cursor.execute("""
                UPDATE attendance_sessions

                SET
                    subject_id=%s,
                    teacher_id=%s,
                    session_date=%s,
                    start_time=%s,
                    end_time=%s,
                    session_status=%s

                WHERE id=%s
            """, (
                subject_id,
                teacher_id,
                session_date,
                start_time,
                end_time if end_time else None,
                session_status,
                id
            ))

            mysql.connection.commit()

            flash(
                "Attendance Session updated successfully!",
                "success"
            )

            return redirect(
                url_for("attendance.index")
            )

        # ==================================================
        # GET SESSION
        # ==================================================

        cursor.execute("""
            SELECT
                id,
                subject_id,
                teacher_id,
                session_date,
                start_time,
                end_time,
                session_status

            FROM attendance_sessions

            WHERE id=%s
        """, (
            id,
        ))

        attendance_session = cursor.fetchone()

        if not attendance_session:

            flash(
                "Attendance session not found.",
                "danger"
            )

            return redirect(
                url_for("attendance.index")
            )

        # ----------------------------------------------
        # Subjects
        # ----------------------------------------------

        cursor.execute("""
            SELECT
                id,
                subject_name
            FROM subjects
            ORDER BY subject_name
        """)

        subjects = cursor.fetchall()

        # ----------------------------------------------
        # Teachers
        # ----------------------------------------------

        cursor.execute("""
            SELECT
                id,
                full_name
            FROM teachers
            ORDER BY full_name
        """)

        teachers = cursor.fetchall()

        return render_template(
            "admin/edit_attendance.html",
            attendance_session=attendance_session,
            subjects=subjects,
            teachers=teachers
        )

    except Exception as e:

        mysql.connection.rollback()

        print(
            "ATTENDANCE SESSION EDIT ERROR:",
            str(e)
        )

        flash(
            f"Error updating attendance session: {e}",
            "danger"
        )

        return redirect(
            url_for("attendance.index")
        )

    finally:

        cursor.close()


# ==========================================================
# Delete Attendance Session
# ==========================================================

@attendance.route(
    "/delete/<int:id>"
)
def delete_session(id):

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    cursor = mysql.connection.cursor()

    try:

        cursor.execute("""
            SELECT id
            FROM attendance_sessions
            WHERE id=%s
            LIMIT 1
        """, (
            id,
        ))

        existing = cursor.fetchone()

        if not existing:

            flash(
                "Attendance session not found.",
                "warning"
            )

            return redirect(
                url_for("attendance.index")
            )

        cursor.execute("""
            DELETE FROM attendance_sessions
            WHERE id=%s
        """, (
            id,
        ))

        mysql.connection.commit()

        flash(
            "Attendance Session deleted successfully!",
            "success"
        )

    except Exception as e:

        mysql.connection.rollback()

        print(
            "ATTENDANCE SESSION DELETE ERROR:",
            str(e)
        )

        flash(
            f"Error deleting attendance session: {e}",
            "danger"
        )

    finally:

        cursor.close()

    return redirect(
        url_for("attendance.index")
    )


# ==========================================================
# Attendance Report
# ==========================================================

@attendance.route("/report")
def report():

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    search = request.args.get(
        "search",
        ""
    ).strip()

    date = request.args.get(
        "date",
        ""
    ).strip()

    method = request.args.get(
        "method",
        ""
    ).strip()

    cursor = mysql.connection.cursor()

    try:

        query = """
            SELECT
                attendance.id,
                students.photo,
                students.student_id,
                students.full_name,
                students.department,
                attendance.attendance_date,
                DAYNAME(attendance.attendance_date)
                    AS day_name,
                attendance.attendance_time,
                attendance.attendance_method,
                attendance.status

            FROM attendance

            INNER JOIN students
                ON attendance.student_id = students.id

            WHERE 1=1
        """

        params = []

        # ----------------------------------------------
        # Search
        # ----------------------------------------------

        if search:

            query += """
                AND
                (
                    students.student_id LIKE %s
                    OR students.full_name LIKE %s
                )
            """

            params.append(
                f"%{search}%"
            )

            params.append(
                f"%{search}%"
            )

        # ----------------------------------------------
        # Date
        # ----------------------------------------------

        if date:

            query += """
                AND attendance.attendance_date=%s
            """

            params.append(date)

        # ----------------------------------------------
        # Method
        # ----------------------------------------------

        if method:

            query += """
                AND attendance.attendance_method=%s
            """

            params.append(method)

        query += """
            ORDER BY
                attendance.attendance_date DESC,
                attendance.attendance_time DESC
        """

        cursor.execute(
            query,
            tuple(params)
        )

        attendance_data = cursor.fetchall()

    except Exception as e:

        print(
            "ATTENDANCE REPORT ERROR:",
            str(e)
        )

        flash(
            f"Unable to load attendance report: {e}",
            "danger"
        )

        attendance_data = []

    finally:

        cursor.close()

    return render_template(
        "admin/attendance_report.html",
        attendance=attendance_data
    )


# ==========================================================
# Export Attendance Report to Excel
# ==========================================================

@attendance.route("/export/excel")
def export_excel():

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    cursor = mysql.connection.cursor()

    try:

        cursor.execute("""
            SELECT
                students.student_id,
                students.full_name,
                students.department,
                attendance.attendance_date,
                DAYNAME(attendance.attendance_date),
                attendance.attendance_time,
                attendance.attendance_method,
                attendance.status

            FROM attendance

            INNER JOIN students
                ON attendance.student_id = students.id

            ORDER BY
                attendance.attendance_date DESC,
                attendance.attendance_time DESC
        """)

        data = cursor.fetchall()

    except Exception as e:

        print(
            "EXCEL EXPORT ERROR:",
            str(e)
        )

        flash(
            f"Unable to export Excel report: {e}",
            "danger"
        )

        return redirect(
            url_for("attendance.report")
        )

    finally:

        cursor.close()

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Attendance Report"

    headers = [
        "Student ID",
        "Full Name",
        "Department",
        "Date",
        "Day",
        "Time",
        "Method",
        "Status"
    ]

    for col, header in enumerate(
        headers,
        start=1
    ):

        sheet.cell(
            row=1,
            column=col
        ).value = header

    row_no = 2

    for row in data:

        for col_no, value in enumerate(
            row,
            start=1
        ):

            sheet.cell(
                row=row_no,
                column=col_no
            ).value = value

        row_no += 1

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="attendance_report.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )


# ==========================================================
# Export Attendance Report to PDF
# ==========================================================

@attendance.route("/export/pdf")
def export_pdf():

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    cursor = mysql.connection.cursor()

    try:

        cursor.execute("""
            SELECT
                students.student_id,
                students.full_name,
                students.department,
                attendance.attendance_date,
                DAYNAME(attendance.attendance_date),
                attendance.attendance_time,
                attendance.attendance_method,
                attendance.status

            FROM attendance

            INNER JOIN students
                ON attendance.student_id = students.id

            ORDER BY
                attendance.attendance_date DESC,
                attendance.attendance_time DESC
        """)

        data = cursor.fetchall()

    except Exception as e:

        print(
            "PDF EXPORT ERROR:",
            str(e)
        )

        flash(
            f"Unable to export PDF report: {e}",
            "danger"
        )

        return redirect(
            url_for("attendance.report")
        )

    finally:

        cursor.close()

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    table_data = [[
        "Student ID",
        "Name",
        "Department",
        "Date",
        "Day",
        "Time",
        "Method",
        "Status"
    ]]

    for row in data:

        table_data.append([
            str(row[0]),
            str(row[1]),
            str(row[2]),
            str(row[3]),
            str(row[4]),
            str(row[5]),
            str(row[6]),
            str(row[7])
        ])

    table = Table(
        table_data,
        colWidths=[
            1.0 * inch,
            1.5 * inch,
            1.2 * inch,
            0.9 * inch,
            0.8 * inch,
            0.9 * inch,
            0.8 * inch,
            0.8 * inch
        ]
    )

    table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.darkblue
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                9
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.black
            ),
            (
                "BACKGROUND",
                (0, 1),
                (-1, -1),
                colors.beige
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, 0),
                8
            )
        ])
    )

    document.build([
        table
    ])

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="attendance_report.pdf",
        mimetype="application/pdf"
    )


# ==========================================================
# Attendance Statistics API
# ==========================================================

@attendance.route("/statistics")
def statistics():

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    cursor = mysql.connection.cursor()

    try:

        # Total Attendance

        cursor.execute("""
            SELECT COUNT(*)
            FROM attendance
        """)

        total_attendance = cursor.fetchone()[0]

        # Today's Attendance

        cursor.execute("""
            SELECT COUNT(*)
            FROM attendance
            WHERE attendance_date = CURDATE()
        """)

        today_attendance = cursor.fetchone()[0]

        # Face Attendance

        cursor.execute("""
            SELECT COUNT(*)
            FROM attendance
            WHERE attendance_method='FACE'
        """)

        face_attendance = cursor.fetchone()[0]

        # QR Attendance

        cursor.execute("""
            SELECT COUNT(*)
            FROM attendance
            WHERE attendance_method='QR'
        """)

        qr_attendance = cursor.fetchone()[0]

        # Manual Attendance

        cursor.execute("""
            SELECT COUNT(*)
            FROM attendance
            WHERE attendance_method='MANUAL'
        """)

        manual_attendance = cursor.fetchone()[0]

    except Exception as e:

        print(
            "ATTENDANCE STATISTICS ERROR:",
            str(e)
        )

        return {
            "total": 0,
            "today": 0,
            "face": 0,
            "qr": 0,
            "manual": 0,
            "error": str(e)
        }

    finally:

        cursor.close()

    return {
        "total": total_attendance,
        "today": today_attendance,
        "face": face_attendance,
        "qr": qr_attendance,
        "manual": manual_attendance
    }


# ==========================================================
# Monthly Attendance Summary
# ==========================================================

@attendance.route("/monthly-summary")
def monthly_summary():

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    cursor = mysql.connection.cursor()

    try:

        cursor.execute("""
            SELECT
                MONTHNAME(attendance_date) AS month,
                COUNT(*) AS total

            FROM attendance

            GROUP BY
                YEAR(attendance_date),
                MONTH(attendance_date)

            ORDER BY
                YEAR(attendance_date),
                MONTH(attendance_date)
        """)

        monthly_data = cursor.fetchall()

    except Exception as e:

        print(
            "MONTHLY ATTENDANCE ERROR:",
            str(e)
        )

        flash(
            f"Unable to load monthly attendance: {e}",
            "danger"
        )

        monthly_data = []

    finally:

        cursor.close()

    return render_template(
        "admin/monthly_attendance.html",
        monthly_data=monthly_data
    )


# ==========================================================
# Department Wise Attendance
# ==========================================================

@attendance.route("/department-summary")
def department_summary():

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    cursor = mysql.connection.cursor()

    try:

        cursor.execute("""
            SELECT
                students.department,
                COUNT(attendance.id)

            FROM attendance

            INNER JOIN students
                ON attendance.student_id = students.id

            GROUP BY students.department

            ORDER BY students.department
        """)

        department_data = cursor.fetchall()

    except Exception as e:

        print(
            "DEPARTMENT ATTENDANCE ERROR:",
            str(e)
        )

        flash(
            f"Unable to load department attendance: {e}",
            "danger"
        )

        department_data = []

    finally:

        cursor.close()

    return render_template(
        "admin/department_attendance.html",
        department_data=department_data
    )


# ==========================================================
# Student Attendance Percentage
# ==========================================================

@attendance.route("/student-percentage")
def student_percentage():

    if "admin_id" not in session:
        return redirect(
            url_for("auth.login")
        )

    cursor = mysql.connection.cursor()

    try:

        cursor.execute("""
            SELECT
                students.student_id,
                students.full_name,
                students.department,
                COUNT(attendance.id)
                    AS total_attendance

            FROM students

            LEFT JOIN attendance
                ON students.id = attendance.student_id

            GROUP BY
                students.id,
                students.student_id,
                students.full_name,
                students.department

            ORDER BY total_attendance DESC
        """)

        student_data = cursor.fetchall()

    except Exception as e:

        print(
            "STUDENT PERCENTAGE ERROR:",
            str(e)
        )

        flash(
            f"Unable to load student attendance: {e}",
            "danger"
        )

        student_data = []

    finally:

        cursor.close()

    return render_template(
        "admin/student_percentage.html",
        students=student_data
    )
